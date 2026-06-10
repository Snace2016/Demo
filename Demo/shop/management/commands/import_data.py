import os
import shutil
from datetime import datetime

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand

from shop.models import (
    AppUser, Category, Manufacturer, Order, OrderItem, OrderStatus,
    PickupPoint, Product, Role, Supplier,
)

ROLE_MAP = {
    'Администратор': Role.ADMIN,
    'Менеджер': Role.MANAGER,
    'Авторизированный клиент': Role.CLIENT,
}

IMAGE_MAP = {
    '1.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_1-da1c9273-1691-4945-b4fc-9fb08a86bf23.png',
    '2.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_2-7e81cffd-7aa9-42f1-927c-ff8739c5bf89.png',
    '3.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_3-818e0358-5fd5-454b-a695-4c430609de92.png',
    '4.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_4-e0c2471d-f65f-42f2-af50-d26cfcc9feb0.png',
    '5.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_5-156db1b0-cb78-4437-a6a2-4f0c8a4a3885.png',
    '6.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_6-c4429155-69c9-4036-b718-83a9f62c73f5.png',
    '7.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_7-c2022a0f-2ca2-4939-840c-1442f5c82146.png',
    '8.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_8-c0892d33-ad5e-40c8-a089-0562cab6cf01.png',
    '9.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_9-d571d909-c2fe-4ee4-9039-13f4e4d2b89f.png',
    '10.jpg': 'c__Users_snace_AppData_Roaming_Cursor_User_workspaceStorage_caf4a529a6799993602f3e42842d6a99_images_10-dac1c997-e79f-422e-989f-f9c07a2fce5a.png',
}


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def copy_images(assets_dir, media_dir):
    products_dir = os.path.join(media_dir, 'products')
    os.makedirs(products_dir, exist_ok=True)
    for target_name, source_name in IMAGE_MAP.items():
        source = os.path.join(assets_dir, source_name)
        if os.path.exists(source):
            shutil.copy2(source, os.path.join(products_dir, target_name))


class Command(BaseCommand):
    help = 'Импорт данных из Excel-файлов'

    def add_arguments(self, parser):
        parser.add_argument(
            '--data-dir',
            default=r'c:\Users\snace\OneDrive\Рабочий стол',
            help='Папка с файлами импорта',
        )

    def handle(self, *args, **options):
        data_dir = options['data_dir']
        assets_dir = settings.BASE_DIR / 'assets'

        self.stdout.write('Очистка базы данных...')
        OrderItem.objects.all().delete()
        Order.objects.all().delete()
        Product.objects.all().delete()
        PickupPoint.objects.all().delete()
        OrderStatus.objects.all().delete()
        AppUser.objects.all().delete()
        Category.objects.all().delete()
        Manufacturer.objects.all().delete()
        Supplier.objects.all().delete()

        copy_images(str(assets_dir), str(settings.MEDIA_ROOT))

        self._import_users(os.path.join(data_dir, 'user_import.xlsx'))
        self._import_pickup_points(os.path.join(data_dir, 'Пункты выдачи_import.xlsx'))
        self._import_products(os.path.join(data_dir, 'Tovar.xlsx'))
        self._import_orders(os.path.join(data_dir, 'Заказ_import.xlsx'))

        self.stdout.write(self.style.SUCCESS('Импорт завершён успешно.'))

    def _import_users(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            AppUser.objects.create(
                role=ROLE_MAP[row[0]],
                full_name=row[1],
                login=row[2],
                password=row[3],
            )
        self.stdout.write(f'  Пользователи: {AppUser.objects.count()}')

    def _import_pickup_points(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = ws.cell(1, 1).value
        if header:
            PickupPoint.objects.create(address=str(header).strip())
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                PickupPoint.objects.create(address=str(row[0]).strip())
        self.stdout.write(f'  Пункты выдачи: {PickupPoint.objects.count()}')

    def _import_products(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            category, _ = Category.objects.get_or_create(name=row[6])
            manufacturer, _ = Manufacturer.objects.get_or_create(name=row[5])
            supplier, _ = Supplier.objects.get_or_create(name=row[4])
            image_name = row[10] or ''
            image_path = f'products/{image_name}' if image_name else ''
            Product.objects.create(
                article=row[0],
                name=row[1],
                unit=row[2],
                price=row[3],
                supplier=supplier,
                manufacturer=manufacturer,
                category=category,
                discount=row[7] or 0,
                quantity=row[8] or 0,
                description=(row[9] or '').strip(),
                image=image_path,
            )
        self.stdout.write(f'  Товары: {Product.objects.count()}')

    def _import_orders(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        pickup_points = list(PickupPoint.objects.order_by('id'))
        clients = {u.full_name: u for u in AppUser.objects.filter(role=Role.CLIENT)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            status_name = (row[7] or 'Новый').strip()
            status, _ = OrderStatus.objects.get_or_create(name=status_name)

            pickup_idx = int(row[4]) - 1
            pickup = pickup_points[pickup_idx] if 0 <= pickup_idx < len(pickup_points) else pickup_points[0]

            client_name = (row[5] or '').strip()
            client = clients.get(client_name)
            if not client:
                client = AppUser.objects.filter(role=Role.CLIENT).first()

            order_date = parse_date(row[2])
            delivery_date = parse_date(row[3])
            if not order_date:
                order_date = datetime(2024, 1, 1).date()
            if not delivery_date:
                delivery_date = datetime(2024, 4, 1).date()

            order = Order.objects.create(
                number=int(row[0]),
                status=status,
                pickup_point=pickup,
                order_date=order_date,
                delivery_date=delivery_date,
                client=client,
                pickup_code=int(row[6]) if row[6] else 900,
            )

            articles_raw = row[1] or ''
            parts = [p.strip() for p in str(articles_raw).split(',') if p.strip()]
            for i in range(0, len(parts), 2):
                if i + 1 >= len(parts):
                    break
                article = parts[i]
                try:
                    qty = int(parts[i + 1])
                except ValueError:
                    continue
                product = Product.objects.filter(article=article).first()
                if product:
                    OrderItem.objects.create(order=order, product=product, quantity=qty)

        self.stdout.write(f'  Заказы: {Order.objects.count()}')
